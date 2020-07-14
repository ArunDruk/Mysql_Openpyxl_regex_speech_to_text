import random
import string
import openpyxl

nums=string.digits
chars=string.ascii_lowercase
Age_list=[]
fname_list=[]
lname_list=[]

num_of_records=10

for i in range(num_of_records):
    # Ages="".join(val for k in range(2) for val in str(random.choice(nums)))
    Ages="".join(random.choice(nums) for k in range(2))
    Age_list.append(Ages)

for i in range(len(Age_list)):
    if (Age_list[i] == '00'):
        Age_list[i]='23'

for i in range(num_of_records):
    fname="".join(random.choice(chars) for k in range(5))
    fname_list.append(fname)

for i in range(num_of_records):
    lname = "".join(random.choice(chars) for k in range(5))
    lname_list.append(lname)

lname_list.insert(0,"Dummy_value")
fname_list.insert(0,"Dummy_value")
Age_list.insert(0,"Dummy_value")

wk=openpyxl.Workbook()
wk.create_sheet("to_write_to_DB")
sh=wk["to_write_to_DB"]
# cell_obj=sh.cell()
title_list=["Sl.No","First Name","Last Name","Age","Mail_ID"]
sh.cell(1,1).value="Name"

for i in range(1,5):
    sh.cell(1,i).value=title_list[i]

for r in range(2,num_of_records+1):
    sh.cell(r,1).value=fname_list[r]
    sh.cell(r,2).value=lname_list[r]
    sh.cell(r,3).value=Age_list[r]
    sh.cell(r,4).value=fname_list[r]+'.'+lname_list[r]+"@techm.com"


wk.save("To_write_to_database1.xlsx")











