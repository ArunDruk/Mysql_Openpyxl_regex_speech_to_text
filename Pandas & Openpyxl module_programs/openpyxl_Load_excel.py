import openpyxl

# Loads the workbook
wk=openpyxl.load_workbook("test1.xlsx")

# print(wk.sheetnames) # This will give sheetnames of the workbook loaded in the form of list

# print(f"Active sheet is: '{wk.active.title}'")
#Creating object for the sheet which we want to work on
sh=wk['Sheet']
# print(sh.title) # Printing the sheet name or the title

##################################### Method 1: To Fetch the value from the cell using sheet object ###############################################################
# print(sh['B4'].value) # Extracting the value from the specific cell,  sh is the sheet object
# print(sh['D9'].value)
######################################## Method 2: To Fetch the value from the cell using Cell object############################################################
c1=sh.cell(4,2)  # Need to enter row * column, c1 is the cell object
print(c1.value)
c1=sh.cell(9,4)  # Need to enter 9th row * 4th column
print(c1.value)
############################################### Method 3: To Fetch the value from the cell using Keyword argument of Cell object###########################################################
c1=sh.cell(column=2, row=4) # Cell object for 
print(c1.value)
print(c1.row)
print(c1.column)
