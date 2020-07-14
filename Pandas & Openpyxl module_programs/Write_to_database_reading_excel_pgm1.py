
import mysql.connector
import pandas as pd
import openpyxl

num_of_records=10

mydb=mysql.connector.connect(user="root",passwd="Feb@2020",host="localhost",database="reading_excel_writing_to_database1")

mycursor=mydb.cursor()

# mycursor.execute("create table techm_employees (Reg_no integer auto_increment primary key, fname varchar(10), lname varchar(10), Age int(3), Email_ID varchar(50))")

# mycursor.execute("show tables")

# sql_query_insert="Insert into techm_employees (fname,lname,Age,Email_ID) Values (%s,%s,%s,%s)"
sql_query_insert="Insert into test_engineers (fname,lname,Age,Email_ID) Values (%s,%s,%s,%s)"
wk=openpyxl.load_workbook("To_write_to_database1.xlsx")
sh=wk["to_write_to_DB"]

for i in range(2,num_of_records+1):
    # val=(sh.cell(i,1).value,sh.cell(i,2).value,sh.cell(i,3).value,sh.cell(i,4).value) # Direct way of doing it but if the number of columns are more, this is not a better way
    values=(",".join(sh.cell(i,j).value for j in range (1,5))) # This method is better if the number of columns are more or less, this just returns string.
    val=values.split(",") # this returns a list or we can do tuple also, to get proper value for each column.
    mycursor.execute(sql_query_insert,val)


mydb.commit()


# df=pd.read_sql_query("show tables",mydb)
# print(df)

