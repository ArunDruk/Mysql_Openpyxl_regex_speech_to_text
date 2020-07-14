import openpyxl

wk1=openpyxl.load_workbook('test1.xlsx')

sh1=wk1['statistics']

# print(sh1['B4'].value)
# c1=sh1.cell(4,2)
# print(c1.value)

############################################################### To get the number of Rows and Columns having data  ###############################################################
num_of_rows=sh1.max_row # To get the max num of rows with data
num_of_columns=sh1.max_column # To get max num of column with data

print(f'The number of rows in "{sh1.title}" are {num_of_rows} and column are {num_of_columns}')

############################################################### Method 1: To access each row and column ###############################################################

# for r in range(1,num_of_rows+1): # have to take complete rows, so rows+1
#     for c in range(1,num_of_columns+1):
#         c1=sh1.cell(r,c)
#         print(f"The row num is {r} and the column num is {c}, the value is {c1.value}")

############################################################### Method 2: To access each row and column ###############################################################

for r in sh1['A1':'D2']:
    for c in r:

        print(f"{c.value}")
##############################################################################################################################