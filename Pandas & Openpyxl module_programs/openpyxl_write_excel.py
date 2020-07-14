import openpyxl

wk1=openpyxl.Workbook() # Creating an Excel or an Workbook
print(wk1.active.title) # To get the active sheet name in the workbook loaded
sh1=wk1.active

sh1.title="Issue_list"  # Renaming the sheet name to the given value
print(wk1.active.title)

sh1['A1'].value="Manheim_file_transfer"  # Inserting value in the particular cell

############################################################### Creating the worksheet ###############################################################
wk1.create_sheet(title='Result1')
###############################################################
sh2=wk1['Result1']
cell_obj1=sh2.cell(4,2)
cell_obj1.value="Inserted_value1"

sh2["C2"].value="Inserted_value22"

############################################################### Removing the worksheet ###############################################################

wk1.remove(wk1["Issue_list"]) # We have to send the sheet object wk1["Issue_list"], not just the name of the sheet

# cell_obj=sh3.cell(5,2)
# print(sh3['B5'].value)
# print(cell_obj.value)

wk1.save("New_wk1.xlsx")  # Saving the workbook, give the path if you need to save it in any directory