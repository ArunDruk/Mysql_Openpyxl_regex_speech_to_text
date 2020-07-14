import openpyxl
import mysql.connector

row=2
# mydb=mysql.connector.connect(user="root",passwd="Feb@2020",host="localhost",database="reading_excel_writing_to_database1")
mydb=mysql.connector.connect(user="root",passwd="Feb@2020",host="localhost",database="world")

mycursor=mydb.cursor()

wk=openpyxl.Workbook()

# wk.create_sheet("populated_from_database")
wk.create_sheet("populated_from_database_world")

sh=wk["populated_from_database_world"]

# mycursor.execute("select * from techm_employees")
mycursor.execute("select * from city where CountryCode='IND'")

# title_list=["Reg_No","First Name","Last Name","Age","Mail_ID"]
title_list=["ID","Name","CountryCode", "District", "Population"]

sh.append(title_list)

for count, i in enumerate(mycursor,start=2):
    sh.append(i)  # To wirte to the excel using List or Tupple to the entire row and not by navigating to each cell.

wk.save("Read_from_table_city.xlsx")